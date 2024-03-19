
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import pandas as pd
import datetime
import pdb 
import re
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select



# Configura as opções do Chrome
webdriver_options = Options()
webdriver_options.add_experimental_option('detach', True)

# Desativa as notificações
prefs = {"profile.default_content_setting_values.notifications" : 2}
webdriver_options.add_experimental_option("prefs", prefs)

# Inicializa o driver do navegador (neste caso, Chrome)
driver = webdriver.Chrome(service=Service('venv\\chromedriver.exe'), options=webdriver_options)


# Espera até que a página tenha terminado de carregar
wait = WebDriverWait(driver, 20)  # Aumenta o tempo limite para 20 segundos

print("Abrindo o site...")
# Abre o site
driver.get("https://ticket.crescer.net")

print("Encontrando os campos de login e senha...")
# Encontra os campos de login e senha
campo_email = driver.find_element(By.XPATH, "//*[@id='username']")
campo_senha = driver.find_element(By.XPATH, "//*[@id='password']")

print("Preenchendo os campos de login e senha...")
# Preenche os campos de login e senha
campo_email.send_keys("claudio.vieira@vivver.com.br")
campo_senha.send_keys("ticket#2023")

print("Clicando no botão de novo ticket...")
# Espera até que o elemento esteja visível na página
elemento = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="login"]/div[4]/button')))
elemento.click()

sleep(10)

# Encontre o elemento //*[@id="content_permanent_TicketOverview"]/div[3]/div[1]/div[1]/div/a[2]/span[1] e clique nele
elemento = driver.find_element(By.XPATH, '//*[@id="content_permanent_TicketOverview"]/div[3]/div[1]/div[1]/div/a[2]/span[1]')
elemento.click()

sleep(3)

# Obter o HTML interno do elemento
elemento = driver.find_element(By.XPATH, '//*[@id="content_permanent_TicketOverview"]/div[3]')
html_interior = elemento.get_attribute('innerHTML')

# Criar um DataFrame com as linhas e colunas da tabela
df = pd.read_html(html_interior, index_col=0)

# Copias para um df os dados da tabela
df = df[0].copy()
# Altera o nome da coluna "#" por ticket
df.rename(columns={'#': 'ticket'}, inplace=True)

# Altera o nome da primeira coluna pora Status
df.rename(columns={'Unnamed: 1': 'Status'}, inplace=True)




# Cria uma lista para ser utilizada em um loop com as linhas da coluna ticket
lista = df['ticket'].tolist()
# Remove linhas cujo conteúdo não seja somente algarismos
lista = [x for x in lista if x.isdigit()]


# Cria um loop com a lista e a cada rodada colalize o elemento "//*[@id="global-search"]" e atribua o conteudo da lista para esse campo

# pdb.set_trace() isso

prioridade = 0
for i in lista:
    sleep(5)
    
    try:
        # Clique no elemento "//*[@id="navigation"]/div[1]/form/div[2]/svg"
        elemento = driver.find_element(By.XPATH, '//*[@id="navigation"]/div[1]/form/div[2]')
        elemento.click()
    except:
        pass    
    
    # Encontre o elemento '//*[@id="navigation"]/div[2]/a[1]/span' e clique nele
    try:
        elemento = driver.find_element(By.XPATH, '//*[@id="navigation"]/div[2]/a[1]/span')
        elemento.click()
    except:
        pass    

    
    # Limpa o campo de busca
    elemento = driver.find_element(By.XPATH, '//*[@id="global-search"]')
    elemento.clear()

    sleep(5)
    elemento = driver.find_element(By.XPATH, '//*[@id="global-search"]')
    
    elemento.send_keys(i)
    sleep(10)

    # Encontre o índice da linha que contém o valor de 'i' na coluna 'ticket'
    index = df[df['ticket'] == i].index[0]
    
    sleep(5)
    # Espera até que o elemento esteja presente na página
    elementos = driver.find_elements(By.XPATH, '//*[contains(@id, "content_permanent_Ticket-")]')

    for index, elemento in enumerate(elementos):
        # Verifique se o elemento é interativo antes de clicar
        if elemento.is_displayed() and elemento.is_enabled():
            elemento.click()
            sleep(5)

            # Salva o texto do elemento na coluna 'descrição' na linha 'i'
            df.loc[index, 'descrição'] = elemento.text

    
    
    
    
    # Encontre todos os elementos que contêm 'Ticket-' em seu ID
    elementos = driver.find_elements(By.XPATH, '//*[contains(@id, "content_permanent_Ticket-")]')

    for elemento in elementos:
        # Extraia o número após 'Ticket-' no ID do elemento
        ticket_number = elemento.get_attribute('id').split('-')[-1]
        
        # Use o número do ticket no XPath
        xpath = f'//*[@id="content_permanent_Ticket-{ticket_number}"]/div/div[2]/div/div[2]/div/div[3]/small/time'
        
        # Espera até que o elemento esteja presente na página
        wait = WebDriverWait(driver, 10)
        elemento = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        
        # Extrai o valor do atributo 'title' do elemento
        title = elemento.get_attribute('title')
        
        # Salva o valor do atributo 'title' na coluna 'descrição' na linha correspondente ao número do ticket
        df.loc[int(ticket_number), ''] = title
    

   

        # Cria um objeto WebDriverWait com um tempo limite de 10 segundos
        wait = WebDriverWait(driver, 10)

        # Use o método until do objeto WebDriverWait para esperar até que o elemento esteja presente
        elemento = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[@id="Ticket_{i}_priorizada"]')))

        # Cria uma instância da classe Select
        select = Select(elemento)

        # Seleciona a opção pelo valor
        select.select_by_value(str(prioridade))

        # prioridade += 1

    # Adicionar as colunas Data_priorizado e hora_atual
    df['Data_priorizado'] = datetime.date.today().strftime('%d/%m/%Y')
                                                
    df['hora_priorizado'] = datetime.datetime.now().strftime("%H:%M:%S")
    
    # Crie uma coluna chamada 'Município' e atribua a ela a primeira ocorrência  entre "[]" da coluna 'Título'
    df['Município'] = df['Título'].str.extract(r'\[(.*?)\]')
       
    # Se Proprietário for = a '-'  preencha com o valor de Proprietário da linha anterior
    df['Proprietário'] = df['Proprietário'].replace('-', method='ffill')

# Remova as linha em que Ticket é igual a NaN
df = df.dropna(subset=['ticket'])

# Se existirem remova linhas não numericas da coluna ticket
df = df[df['ticket'].str.isnumeric()]

# format ticket e Priorizada to int
df['ticket'] = df['ticket'].astype(int)
df['Priorizada'] = df['Priorizada'].astype(int)


df_com_prioridade = df



# Lê a planilha existente em um DataFrame
try:
    df_existente = pd.read_excel(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx')
    
    

except FileNotFoundError:
    df_existente = pd.DataFrame()
    


# Adiciona os novos registros ao DataFrame existente
df_final = pd.concat([df_existente, df])

# Converter a coluna 'ticket' para string e preencher os valores NaN com uma string vazia
df_final['ticket'] = df_final['ticket'].fillna('').astype(str)

# Remover as linhas da coluna 'ticket' que não são numéricas
df_final = df_final[df_final['ticket'].str.isnumeric()]

# Escreve o DataFrame de volta para a planilha
df_final.to_excel(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', index=False)

# Agrupar as linhas que têm as mesmas informações nas colunas "ticket" e "Data_priorizado"
df_final_agrupado = df_final.groupby(['ticket', 'Data_priorizado']).first().reset_index()

# Criar uma nova aba no arquivo priorizadas_semana.xlsx e escrever o DataFrame df_final_agrupado
with pd.ExcelWriter(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', mode='a') as writer:
    df_final_agrupado.to_excel(writer, sheet_name='agrupado', index=False)


with pd.ExcelWriter(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_com_prioridade.to_excel(writer, sheet_name='com_prioridade', index=False)
    
# Contar a quantidade de vezes que cada ticket aparece
contagem_tickets = df_final['ticket'].value_counts()

# Transformar a Series em DataFrame e resetar o índice
contagem_tickets = contagem_tickets.reset_index()
contagem_tickets.columns = ['ticket', 'Dias_priorizados']

# Criar uma nova aba no arquivo priorizadas_semana.xlsx e escrever o DataFrame contagem_tickets
with pd.ExcelWriter(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', mode='a') as writer:
    contagem_tickets.to_excel(writer, sheet_name='contagem_tickets', index=False)
    
# Ler os dados da aba 'agrupado'
df_agrupado = pd.read_excel(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', sheet_name='agrupado')
# Contar a quantidade de vezes que cada ticket aparece

contagem_tickets_agrupado = df_agrupado['ticket'].value_counts()

# Transformar a Series em DataFrame e resetar o índice
contagem_tickets_agrupado = contagem_tickets_agrupado.reset_index()
contagem_tickets_agrupado.columns = ['ticket', 'Dias_priorizado']

# Adicionar as fórmulas do Excel como novas colunas
contagem_tickets_agrupado = contagem_tickets_agrupado.assign(
    Município="",
    Título=""
)

#Criar uma aba no arquivo priorizadas_semana.xlsx e escrever o DataFrame df_com_prioridade
with pd.ExcelWriter('priorizadas_semana.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_com_prioridade.to_excel(writer, sheet_name='df_com_prioridade', index=False)


# Criar uma nova aba no arquivo priorizadas_semana.xlsx e escrever o DataFrame contagem_tickets_agrupado
with pd.ExcelWriter(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    contagem_tickets_agrupado.to_excel(writer, sheet_name='contagem_tickets_agrupado', index=False)

# Abrir o arquivo Excel com openpyxl
wb = load_workbook(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx')

# Selecionar a aba 'contagem_tickets_agrupado'
sheet = wb['contagem_tickets_agrupado']

# Escrever as fórmulas nas células C2 até E51
for i in range(2, 52):
    sheet[f'C{i}'] = f'=IFERROR(VLOOKUP(A{i},Sheet1!A1:L800,10,FALSE), "Retornados")'
    sheet[f'D{i}'] = f'=IFERROR(VLOOKUP(A{i},Sheet1!A1:L900,2,FALSE), "Retornados")'
    
    
# Salvar o arquivo Excel
wb.save(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx')

# Esse trecho abre a planilha de Tickts e insere a quantidade de dias que um determinado Ticket foi priorizado buscando essa informação na planilha Priorizados_Semana
wb = load_workbook(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx')

# Ler todas as abas do arquivo 'priorizadas_semana.xlsx'
priorizadas_semana = pd.read_excel(r'C:\Users\Regina Casoti\Documents\Vivver\priorizadas_semana.xlsx', sheet_name=None)

# Carregar o arquivo 'Tickets_Vivver.xlsx'
book = load_workbook(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx')

# Remover todas as abas exceto a aba 'Todos'
for sheet in book.sheetnames:
    if sheet != 'Todos':
        del book[sheet]

# Salvar as alterações
book.save(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx')

# Abrir o arquivo 'Tickets_Vivver.xlsx' com o ExcelWriter
with pd.ExcelWriter(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx', engine='openpyxl', mode='a') as writer:
    # Iterar sobre todas as abas em 'priorizadas_semana'
    for sheet_name, df in priorizadas_semana.items():
        # Escrever cada aba no arquivo 'Tickets_Vivver.xlsx'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Agora, 'Tickets_Vivver.xlsx' deve conter todas as abas de 'priorizadas_semana.xlsx'


# Carregar a planilha
wb = load_workbook(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx')

# Selecionar a aba "Todos"
sheet = wb['Todos']

# Escrever a fórmula nas células I2 até I100 e J2 até J100
for i in range(2, 101):
    sheet[f'I{i}'] = f'=IFERROR(VLOOKUP(B{i}, com_prioridade!$A$1:$H$11, 4, FALSE), "")'

# Escrever a fórmula nas células I2 até I100 e J2 até J100
for i in range(2, 101):
    sheet[f'J{i}'] = f'=IFERROR(VLOOKUP(B{i}, contagem_tickets_agrupado!$A$1:$B$30, 2, FALSE), "")'


# Inserir a fórmula na coluna 'FeedBack' (coluna 'M') da linha 8 até a linha 30
for i in range(8, 65):
    sheet[f'M{i}'] = f'=IFERROR(VLOOKUP(B{i},Retornados!B$1:H$9,3,FALSE), "")'


# Inserir a fórmula na coluna 'Pendentes' (coluna 'L') da linha 8 até a linha 30
for i in range(8, 65):
    sheet[f'L{i}'] = f'=IFERROR(VLOOKUP(B{i},Pendentes!A$1:G$4030, 4, FALSE), "")'


# Inserir a fórmula na coluna 'Fechado' (coluna 'K') da linha 8 até a linha 30
for i in range(8, 65):
    sheet[f'K{i}'] = f'=IFERROR(VLOOKUP(B{i},Fechados!B$2694:L$4001, 3, FALSE), "")'





# Salvar o arquivo Excel
wb.save(r'C:\Users\Regina Casoti\Documents\Vivver\Tickets_Vivver.xlsx')