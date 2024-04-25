from flask import Flask, render_template, request
import pandas as pd
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime
import time

app = Flask(__name__)

# O código existente foi movido para funções que podem ser chamadas quando necessário
def make_request(url):
    try:
        response = requests.get(url, verify=False)  # Ignore SSL certificate errors
        if response.status_code == 200:
            return response
    except requests.exceptions.RequestException as e:
        print(f"Erro ao fazer a solicitação para {url}: {e}")
    return None

def update_dataframe(df, index, version, now):
    if df.loc[index, 'Versão_Anterior'] != df.loc[index, 'Versão_Atual']:
        df.loc[index, 'Versão_Anterior'] = df.loc[index, 'Versão_Atual']
    else:
        df.loc[index, 'Data_atualização _Anterior'] = now.strftime("%d/%m/%Y %H:%M:%S")

def apply_styles_to_excel(df, wb):
    sheet = wb.active

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for row in sheet.iter_rows(min_row=2):
        if row[1].value != row[2].value:  # Compare 'Versão_Anterior' and 'Versão_Atual'
            for cell in row:
                cell.fill = green_fill
        else:
            for cell in row:
                cell.fill = yellow_fill

    updated_column_index = df.columns.get_loc('Atualizada')

    for row in sheet.iter_rows(min_row=2):
        if row[updated_column_index].value == "Versão Mais Recente":
            row[updated_column_index].fill = blue_fill

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        user_version = request.form.get('version')
        now = datetime.now()
        df = pd.read_excel('municipio_versao.xlsx')

        if 'Data_atualização _Anterior' not in df.columns:
            df['Data_atualização _Anterior'] = None
        df['Data_atualização _Anterior'] = df['Data_atualização _Anterior'].astype(str)

        with requests.Session() as s:
            for index, row in df.iterrows():
                url = f"https://www.{row['Município']}.vivver.com/login"
                response = make_request(url)
                if response is None:
                    url = f"https://{row['Município']}.vivver.com/login"
                    response = make_request(url)
                    if response is None:
                        continue

                time.sleep(3)
                soup = BeautifulSoup(response.text, 'html.parser')
                version_element = soup.find(class_='vmx-login-versao')

                if version_element is not None:
                    version_text = version_element.get_text(strip=True)
                    version = version_text.split(':')[-1].strip()
                    update_dataframe(df, index, version, now)

        df.to_excel('municipio_versao.xlsx', index=False)

        wb = load_workbook('municipio_versao.xlsx')
        apply_styles_to_excel(df, wb)
        wb.save('municipio_versao.xlsx')

    return render_template('index.html')

@app.route('/data')
def data():
    df = pd.read_excel('municipio_versao.xlsx')
    return df.to_html()

if __name__ == '__main__':
    app.run(debug=True)