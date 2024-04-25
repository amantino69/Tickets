import pandas as pd
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime
import time
from packaging import version

# Ler o arquivo Excel
df = pd.read_excel('municipio_versao.xlsx')

# Convert 'Data_atualização_Anterior' columns to string
if 'Data_atualização_Anterior' not in df.columns:
    df['Data_atualização_Anterior'] = None
df['Data_atualização_Anterior'] = df['Data_atualização_Anterior'].astype(str)

# Get the version from the user
user_version = input("Digite o número da versão atual: ")

# Get the current date and time
now = datetime.now()

# Iterar sobre todas as linhas do DataFrame
for index, row in df.iterrows():
    ...
    try:
        # Construir a URL com 'www'
        url = f"https://www.{row['Município']}.vivver.com/login"
        # Fazer a solicitação HTTP
        response = requests.get(url, verify=False)  # Ignore SSL certificate errors
    except requests.exceptions.RequestException:
        try:
            # Construir a URL sem 'www'
            url = f"https://{row['Município']}.vivver.com/login"
            # Fazer a solicitação HTTP
            response = requests.get(url, verify=False)  # Ignore SSL certificate errors
        except requests.exceptions.RequestException as e:
            print(f"Erro ao fazer a solicitação para {url}: {e}")
            continue  # Skip to the next iteration of the loop

    try:
        # Fazer a solicitação HTTP
        response = requests.get(url, verify=False)  # Ignore SSL certificate errors

        # Verificar se a solicitação foi bem-sucedida
        if response.status_code == 200:
            # Aguardar 3 segundos para dar tempo da página carregar
            time.sleep(3)

            # Analisar o HTML da página
            soup = BeautifulSoup(response.text, 'html.parser')

            # Encontrar o elemento com a classe 'vmx-login-versao'
            version_element = soup.find(class_='vmx-login-versao')

            # Verificar se o elemento foi encontrado
            if version_element is not None:
                # Extrair o texto do elemento
                version_text = version_element.get_text(strip=True)

                # Extrair a versão do texto
                version = version_text.split(':')[-1].strip()
                

                # Update 'Data_Atualizaçã_Atual' with the current date and time
                # df.loc[index, 'Data_Atualizaçã_Atual'] = now.strftime("%d/%m/%Y %H:%M:%S")

                # Compare 'Versão_Anterior' and 'Versão_Atual'
                if row[1] == version:
                    # Update 'Versão_Anterior' with the value of 'Versão_Atual'
                    #df.loc[index, 'Versão_Anterior'] = row['Versão_Atual']
                    # Keep the 'Data_atualização_Anterior' date
                    pass
                else:
                    # Update 'Data_atualização_Anterior' with the current date and time
                    df.loc[index, 'Versão_Atual'] = version
                    df.loc[index, 'Data_Atualizaçã_Atual'] = now.strftime("%d/%m/%Y %H:%M:%S")


                # Update 'Versão_Atual' with the new version

    except (requests.exceptions.RequestException, requests.exceptions.SSLError) as e:
        print(f"Erro ao fazer a solicitação para {url}: {e}")

# Iterate over the rows
for index, row in df.iterrows():
    # Compare 'Versão_Atual' with the user's version
    if str(row['Versão_Atual']) == str(user_version):
        # Update 'Atualizada' with "Versão Mais Recente"
        df.loc[index, 'Atualizada'] = "Versão Mais Recente"

# Save the DataFrame to the same Excel file
df.to_excel('municipio_versao.xlsx', index=False)

# Load the workbook and select the first sheet
wb = load_workbook('municipio_versao.xlsx')
sheet = wb.active

green_fill = PatternFill(start_color="00FF00",
                         end_color="00FF00",
                         fill_type="solid")

yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00",
                          fill_type="solid")

# Define a border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Define a fill style for "Atualizada" cells
blue_fill = PatternFill(start_color="ADD8E6",
                        end_color="ADD8E6",
                        fill_type="solid")

# Iterate over the rows
for row in sheet.iter_rows(min_row=2):
    if row[1].value != row[3].value:  # Compare 'Versão_Anterior' and 'Versão_Atual'
        print("row[1]:", row[1].value)
        print("row[3]:", row[3].value)
        for cell in row:
            cell.fill = green_fill
    else:
        for cell in row:
            cell.fill = yellow_fill

# Get the index of 'Atualizada' column
updated_column_index = df.columns.get_loc('Atualizada')

# Iterate over the rows
for row in sheet.iter_rows(min_row=2):
    if row[updated_column_index].value == "Versão Mais Recente":
        row[updated_column_index].fill = blue_fill

# Iterate over the columns
for column in sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Iterate over the rows and columns
for row in sheet.iter_rows():
    for cell in row:
        # Apply border to all cells
        cell.border = thin_border

# Save the workbook
wb.save('municipio_versao.xlsx')